#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
股价记录模块
将每次查询的股价记录到Excel文件
每个股票一个Sheet，便于分析
"""

import os
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook, load_workbook


class PriceLogger:
    """股价记录器 - 每个股票一个Sheet"""
    
    def __init__(self, excel_path: str = None):
        """初始化
        
        Args:
            excel_path: Excel文件路径，默认在项目根目录
        """
        if excel_path is None:
            project_dir = Path(__file__).parent.parent
            excel_path = project_dir / "股价记录.xlsx"
        
        self.excel_path = str(excel_path)
        self._init_excel()
    
    def _init_excel(self):
        """初始化Excel文件"""
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "说明"
            ws.append(["股票代码", "说明"])
            ws.append(["示例: 601166", "每个股票一个Sheet"])
            wb.save(self.excel_path)
            print(f"✅ 创建股价记录文件: {self.excel_path}")
    
    def _get_or_create_sheet(self, stock_code: str, stock_name: str = ""):
        """获取或创建股票的Sheet
        
        Args:
            stock_code: 股票代码
            stock_name: 股票名称（可选）
        """
        wb = load_workbook(self.excel_path)
        
        # 检查Sheet是否已存在
        if stock_code in wb.sheetnames:
            return wb[stock_code]
        
        # 创建新Sheet
        ws = wb.create_sheet(title=stock_code)
        
        # 写入表头
        headers = ["时间", "当前价", "涨跌", "涨跌幅(%)", "开盘", "最高", "最低", "昨收", "成交量", "成交额(元)", "备注"]
        ws.append(headers)
        
        # 写入股票信息
        ws.insert_rows(1)
        ws.insert_rows(1)
        ws['A1'] = f"股票代码: {stock_code}"
        ws['A2'] = f"股票名称: {stock_name}"
        
        # 冻结表头
        ws.freeze_panes = "A4"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20  # 时间
        ws.column_dimensions['B'].width = 12  # 当前价
        ws.column_dimensions['C'].width = 10  # 涨跌
        ws.column_dimensions['D'].width = 12  # 涨跌幅
        ws.column_dimensions['E'].width = 10  # 开盘
        ws.column_dimensions['F'].width = 10  # 最高
        ws.column_dimensions['G'].width = 10  # 最低
        ws.column_dimensions['H'].width = 10  # 昨收
        ws.column_dimensions['I'].width = 15  # 成交量
        ws.column_dimensions['J'].width = 15  # 成交额
        ws.column_dimensions['K'].width = 20  # 备注
        
        wb.save(self.excel_path)
        print(f"✅ 创建新股票Sheet: {stock_code} ({stock_name})")
        
        return wb[stock_code]
    
    def log_price(self, price_data: dict, note: str = ""):
        """记录股价
        
        Args:
            price_data: 股价数据字典
            note: 备注
        """
        try:
            stock_code = price_data.get("code", "")
            stock_name = price_data.get("name", "")
            
            if not stock_code:
                print("❌ 股票代码不能为空")
                return False
            
            # 每次都重新加载工作簿
            wb = load_workbook(self.excel_path)
            
            # 检查Sheet是否已存在（使用股票名称作为Sheet名）
            sheet_name = stock_name if stock_name else stock_code
            if sheet_name not in wb.sheetnames:
                # 创建新Sheet
                ws = wb.create_sheet(title=sheet_name)
                
                # 写入股票信息
                ws['A1'] = f"股票代码: {stock_code}"
                ws['A2'] = f"股票名称: {stock_name}"
                
                # 写入表头
                headers = ["时间", "当前价", "涨跌", "涨跌幅(%)", "开盘", "最高", "最低", "昨收", "成交量", "成交额(元)", "备注"]
                ws.append(headers)
                
                # 冻结表头
                ws.freeze_panes = "A4"
                
                # 设置列宽
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 10
                ws.column_dimensions['F'].width = 10
                ws.column_dimensions['G'].width = 10
                ws.column_dimensions['H'].width = 10
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 20
                
                print(f"✅ 创建新股票Sheet: {stock_code} ({stock_name})")
            else:
                ws = wb[stock_code]
            
            # 准备数据
            row_data = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                price_data.get("price", 0),
                price_data.get("change", 0),
                price_data.get("change_pct", 0),
                price_data.get("open", 0),
                price_data.get("high", 0),
                price_data.get("low", 0),
                price_data.get("prev", 0),
                price_data.get("volume", 0),
                price_data.get("amount", 0),
                note
            ]
            
            # 追加数据
            ws.append(row_data)
            
            # 保存
            wb.save(self.excel_path)
            
            print(f"✅ 股价已记录: {stock_name} {price_data.get('price')}元")
            return True
            
        except Exception as e:
            print(f"❌ 记录股价失败: {e}")
            return False
    
    def log_multiple(self, price_list: list):
        """批量记录多只股票"""
        for price_data in price_list:
            self.log_price(price_data)


# 测试
if __name__ == "__main__":
    logger = PriceLogger()
    
    # 测试记录
    test_data = {
        "code": "601166",
        "name": "兴业银行",
        "price": 18.41,
        "change": -0.43,
        "change_pct": 1.08,
        "open": 18.55,
        "high": 18.59,
        "low": 18.39,
        "prev": 18.49,
        "volume": 493751,
        "amount": 912205941
    }
    
    logger.log_price(test_data, "测试记录")
